"""
Bot de Telegram — Liquidacion de Gastos COVISIAN ESPANA
Multi-usuario con registro. Una fila por gasto.

Categorias:
  - Kilometraje: ida+vuelta desde Covisian, con origen y destino
  - Parking: parkimetro o parking privado
  - Taxi: Uber, Cabify, taxi
  - Comidas: desayunos, cafes, comidas
  - Otros: resto

Requisitos:
  pip install python-telegram-bot anthropic openpyxl pillow --user
"""

import os
import base64
import json
import logging
from datetime import datetime
from pathlib import Path

from telegram import Update
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    filters, ContextTypes, ConversationHandler
)
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TELEGRAM_TOKEN    = os.getenv("TELEGRAM_TOKEN", "TU_TOKEN_AQUI")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "TU_KEY_AQUI")

EMPRESA        = "COVISIAN ESPANA"
CODIGO_CAMPANA = "CAMP 3495"
PRECIO_KM      = 0.12
ORIGEN_FIJO    = "Av. de la Albufera 319, Madrid"
USERS_FILE     = "usuarios.json"

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)
client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

HOY = datetime.now().strftime("%d/%m/%Y")
NOMBRE_STATE, DEPTO_STATE = range(2)


# ═══════════════════════════════════════════════════════════
#  USUARIOS
# ═══════════════════════════════════════════════════════════

def load_users():
    if Path(USERS_FILE).exists():
        with open(USERS_FILE) as f:
            return json.load(f)
    return {}

def save_users(users):
    with open(USERS_FILE, "w") as f:
        json.dump(users, f, ensure_ascii=False, indent=2)

def get_user(tid):
    return load_users().get(str(tid))

def register_user(tid, nombre, departamento):
    users = load_users()
    users[str(tid)] = {"nombre": nombre, "departamento": departamento}
    save_users(users)


# ═══════════════════════════════════════════════════════════
#  EXCEL
# ═══════════════════════════════════════════════════════════

C_TITULO   = "1F3864"
C_CABECERA = "2E75B6"
C_FILA_PAR = "EBF3FB"
C_BORDE    = "AAAAAA"

COLS = [1,2,3,4,5,6,7,8,9,10,11,12,13]

def _border():
    s = Side(style="thin", color=C_BORDE)
    return Border(left=s, right=s, top=s, bottom=s)

def get_filename(nombre):
    n = datetime.now()
    return f"liquidacion_{nombre.replace(' ','_')}_{n.year}_{n.month:02d}.xlsx"

def _build_template(ws, nombre, departamento):
    ws.merge_cells("G1:N2")
    t = ws["G1"]
    t.value = f"LIQUIDACION DE GASTOS {EMPRESA}"
    t.font = Font(bold=True, size=13, name="Arial", color=C_TITULO)
    t.alignment = Alignment(horizontal="center", vertical="center")

    ws["G4"] = "EMPLEADO:";     ws["H4"] = nombre
    ws["G5"] = "DEPARTAMENTO:"; ws["H5"] = departamento
    ws["G6"] = "FECHA:";        ws["H6"] = HOY
    for r in [4,5,6]:
        ws.cell(row=r,column=7).font = Font(bold=True,name="Arial",size=9)
        ws.cell(row=r,column=8).font = Font(name="Arial",size=9)

    ws["E4"] = "Kilometro coche empresa"; ws["F4"] = 0.12
    ws["F4"].number_format = "0.00"

    h8 = ["FECHA","CODIGO\nCAMPANA","CONCEPTO","ORIGEN","DESTINO",
          "KILOMETRAJE","","","PARKING","TAXIS","COMIDAS","OTROS","TOTAL"]
    h9 = ["","","","","","KMS.","EUR/KM","IMPORTE","","","","",""]

    for i,col in enumerate(COLS):
        for row,val in [(8,h8[i]),(9,h9[i])]:
            c = ws.cell(row=row,column=col,value=val)
            c.font = Font(bold=True,color="FFFFFF",name="Arial",size=8)
            c.fill = PatternFill("solid",fgColor=C_CABECERA)
            c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
            c.border = _border()

    ws.merge_cells("F8:H8")
    widths = [10,12,18,22,22,7,7,9,9,9,9,9,9]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[8].height = 20
    ws.row_dimensions[9].height = 16
    ws.freeze_panes = "A10"

def get_or_create_wb(filename, nombre, departamento):
    if Path(filename).exists():
        wb = openpyxl.load_workbook(filename)
        return wb, wb.active
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liquidacion"
    _build_template(ws, nombre, departamento)
    wb.save(filename)
    return wb, ws

def _next_row(ws):
    row = 10
    while ws.cell(row=row,column=1).value is not None:
        row += 1
    return row

def append_expense(user, data):
    filename = get_filename(user["nombre"])
    wb, ws   = get_or_create_wb(filename, user["nombre"], user["departamento"])
    row      = _next_row(ws)
    fill     = C_FILA_PAR if row % 2 == 0 else "FFFFFF"

    values = {
        1:  data.get("fecha", HOY),
        2:  CODIGO_CAMPANA,
        3:  data.get("concepto",""),
        4:  data.get("origen",""),
        5:  data.get("destino",""),
        6:  data.get("kms"),
        7:  data.get("precio_km"),
        8:  data.get("importe_km"),
        9:  data.get("parking"),
        10: data.get("taxis"),
        11: data.get("comidas"),
        12: data.get("otros"),
        13: data.get("total"),
    }

    for col,val in values.items():
        c = ws.cell(row=row,column=col,value=val)
        c.fill = PatternFill("solid",fgColor=fill)
        c.border = _border()
        c.font = Font(name="Arial",size=9)
        c.alignment = Alignment(vertical="center")
        if col in [8,9,10,11,12,13]: c.number_format = '#,##0.00 EUR'
        elif col == 7: c.number_format = '0.00 EUR'
        elif col == 6: c.number_format = '0.00'

    ws.row_dimensions[row].height = 16

    tr = row + 2
    ws.cell(row=tr,column=3,value="TOTAL").font = Font(bold=True,name="Arial",size=9)
    for col in [8,9,10,11,12,13]:
        cl = get_column_letter(col)
        c  = ws.cell(row=tr,column=col,value=f"=SUM({cl}10:{cl}{row})")
        c.font = Font(bold=True,name="Arial",size=9)
        c.fill = PatternFill("solid",fgColor="D6E4F7")
        c.border = _border()
        c.number_format = '#,##0.00 EUR'

    wb.save(filename)
    return row, filename


# ═══════════════════════════════════════════════════════════
#  CLAUDE
# ═══════════════════════════════════════════════════════════

PROMPT_TICKET = """Analiza este ticket. El pie de foto indica el cliente.
Clasifica en UNA sola categoria: parking, taxi, comida, otros.
- parking: parkimetro, parking privado, hora de aparcamiento
- taxi: Uber, Cabify, taxi, transfer
- comida: desayuno, cafe, comida, cena, restaurante
- otros: cualquier otra cosa

Devuelve SOLO este JSON sin backticks:
{
  "fecha": "DD/MM/YYYY o null",
  "cliente": "nombre del cliente del pie de foto",
  "categoria": "parking" o "taxi" o "comida" o "otros",
  "total": numero o null
}
Pie de foto: "{CAPTION}" """

PROMPT_KM = f"""El usuario ha hecho un desplazamiento en coche desde {ORIGEN_FIJO}.
Extrae cliente y direccion exacta de destino.
Calcula km de IDA Y VUELTA entre {ORIGEN_FIJO} y ese destino en Madrid.
Devuelve SOLO este JSON sin backticks:
{{
  "fecha": "DD/MM/YYYY (hoy es {HOY} si no se indica)",
  "cliente": "nombre del cliente mencionado",
  "destino": "direccion exacta de destino",
  "kms_ida": numero km de ida,
  "kms_total": numero km ida y vuelta
}}
Mensaje: "{{MENSAJE}}" """


def extract_from_image(image_bytes, caption=""):
    b64 = base64.standard_b64encode(image_bytes).decode()
    prompt = PROMPT_TICKET.replace("{CAPTION}", caption or "sin pie de foto")
    msg = client.messages.create(
        model="claude-opus-4-20250514", max_tokens=300,
        messages=[{"role":"user","content":[
            {"type":"image","source":{"type":"base64","media_type":"image/jpeg","data":b64}},
            {"type":"text","text":prompt}
        ]}]
    )
    raw = msg.content[0].text.strip().replace("```json","").replace("```","").strip()
    return json.loads(raw)

def extract_km(texto):
    prompt = PROMPT_KM.replace("{MENSAJE}", texto)
    msg = client.messages.create(
        model="claude-opus-4-20250514", max_tokens=300,
        messages=[{"role":"user","content":prompt}]
    )
    raw = msg.content[0].text.strip().replace("```json","").replace("```","").strip()
    return json.loads(raw)

def is_km(texto):
    kw = ["he ido","fui a","visite","visit","estuve en","reunion en",
          "he visitado","desplazado","desplazamiento","km","kilometro","coche"]
    return any(k in texto.lower() for k in kw)


# ═══════════════════════════════════════════════════════════
#  HANDLERS
# ═══════════════════════════════════════════════════════════

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tid  = update.effective_user.id
    user = get_user(tid)
    if user:
        await update.message.reply_text(
            f"Hola *{user['nombre']}*, ya estas registrado.\n\n"
            "Foto de ticket con cliente en el pie de foto\n"
            "O escribe: _He ido a GLS, Calle Ejemplo 23_\n\n"
            "/excel — tu hoja del mes\n"
            "/resumen — tus totales",
            parse_mode="Markdown"
        )
        return ConversationHandler.END
    await update.message.reply_text(
        "Bienvenido al Bot de Gastos COVISIAN\n\n"
        "Primera vez. Necesito registrarte.\n\n"
        "Cual es tu nombre completo?"
    )
    return NOMBRE_STATE

async def get_nombre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["nombre"] = update.message.text.strip()
    await update.message.reply_text("Y tu departamento?")
    return DEPTO_STATE

async def get_depto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nombre = context.user_data["nombre"]
    depto  = update.message.text.strip()
    register_user(update.effective_user.id, nombre, depto)
    await update.message.reply_text(
        f"Registrado correctamente\n\n"
        f"Nombre: {nombre}\n"
        f"Departamento: {depto}\n\n"
        "Ya puedes mandarme fotos o escribir desplazamientos."
    )
    return ConversationHandler.END

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tid  = update.effective_user.id
    user = get_user(tid)
    if not user:
        await update.message.reply_text("Primero registrate con /start")
        return

    msg = await update.message.reply_text("Analizando ticket...")
    try:
        photo       = update.message.photo[-1]
        file        = await context.bot.get_file(photo.file_id)
        image_bytes = bytes(await file.download_as_bytearray())
        caption     = update.message.caption or ""

        data      = extract_from_image(image_bytes, caption)
        categoria = data.get("categoria","otros")
        total     = data.get("total")
        fecha     = data.get("fecha") or datetime.now().strftime("%d/%m/%Y")
        cliente   = data.get("cliente") or caption or "Sin cliente"

        row_data = {
            "fecha":   fecha,
            "concepto": cliente,
            "parking": total if categoria == "parking" else None,
            "taxis":   total if categoria == "taxi"    else None,
            "comidas": total if categoria == "comida"  else None,
            "otros":   total if categoria == "otros"   else None,
            "total":   total,
        }

        row, _ = append_expense(user, row_data)

        emoji = {"parking":"🅿️","taxi":"🚕","comida":"🍽️","otros":"📦"}.get(categoria,"📦")
        await msg.edit_text(
            f"Registrado en fila {row}\n\n"
            f"Fecha: {fecha}\n"
            f"Cliente: {cliente}\n"
            f"{emoji} {categoria.upper()}\n"
            f"Total: {total} EUR"
        )
    except Exception as e:
        logger.exception("Error foto")
        await msg.edit_text(f"Error: {e}")

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tid  = update.effective_user.id
    user = get_user(tid)
    if not user:
        await update.message.reply_text("Primero registrate con /start")
        return

    texto = update.message.text.strip()

    if not is_km(texto):
        await update.message.reply_text(
            "No he detectado un desplazamiento.\n\n"
            "Escribe por ejemplo:\n"
            "He ido a GLS, Calle Fuencarral 123, Madrid\n\n"
            "O manda una foto del ticket con el cliente en el pie de foto."
        )
        return

    msg = await update.message.reply_text("Calculando km ida y vuelta...")
    try:
        data    = extract_km(texto)
        kms     = float(data.get("kms_total",0) or 0)
        kms_ida = float(data.get("kms_ida",0) or 0)
        importe = round(kms * PRECIO_KM, 2)
        fecha   = data.get("fecha") or datetime.now().strftime("%d/%m/%Y")
        destino = data.get("destino","")
        cliente = data.get("cliente","")

        row_data = {
            "fecha":      fecha,
            "concepto":   cliente,
            "origen":     ORIGEN_FIJO,
            "destino":    destino,
            "kms":        kms,
            "precio_km":  PRECIO_KM,
            "importe_km": importe,
            "total":      importe,
        }

        row, _ = append_expense(user, row_data)

        await msg.edit_text(
            f"Desplazamiento registrado en fila {row}\n\n"
            f"Fecha: {fecha}\n"
            f"Cliente: {cliente}\n"
            f"Covisian -> {destino} -> Covisian\n"
            f"{kms_ida} km ida x 2 = {kms} km\n"
            f"{kms} km x {PRECIO_KM} EUR = {importe} EUR"
        )
    except Exception as e:
        logger.exception("Error texto")
        await msg.edit_text(f"Error: {e}")

async def send_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tid  = update.effective_user.id
    user = get_user(tid)
    if not user:
        await update.message.reply_text("Primero registrate con /start")
        return
    filename = get_filename(user["nombre"])
    if not Path(filename).exists():
        await update.message.reply_text("Sin gastos registrados este mes.")
        return
    await update.message.reply_document(
        document=open(filename,"rb"),
        filename=filename,
        caption=f"Liquidacion {datetime.now().strftime('%B %Y')} - {user['nombre']}"
    )

async def resumen(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tid  = update.effective_user.id
    user = get_user(tid)
    if not user:
        await update.message.reply_text("Primero registrate con /start")
        return
    filename = get_filename(user["nombre"])
    if not Path(filename).exists():
        await update.message.reply_text("Sin gastos este mes.")
        return

    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.active
    parking = taxis = comidas = otros = kms_total = filas = 0.0

    for row in ws.iter_rows(min_row=10, values_only=True):
        if not any(row): continue
        filas    += 1
        kms_total += float(row[5] or 0)
        parking  += float(row[8] or 0)
        taxis    += float(row[9] or 0)
        comidas  += float(row[10] or 0)
        otros    += float(row[11] or 0)

    total = parking + taxis + comidas + otros
    await update.message.reply_text(
        f"Resumen {datetime.now().strftime('%B %Y')}\n"
        f"Empleado: {user['nombre']}\n\n"
        f"Registros: {int(filas)}\n"
        f"Km totales: {kms_total:.1f} km\n"
        f"Parking: {parking:.2f} EUR\n"
        f"Taxis: {taxis:.2f} EUR\n"
        f"Comidas: {comidas:.2f} EUR\n"
        f"Otros: {otros:.2f} EUR\n"
        f"TOTAL: {total:.2f} EUR"
    )


# ═══════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════

def main():
    app = ApplicationBuilder().token(TELEGRAM_TOKEN).build()

    conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            NOMBRE_STATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_nombre)],
            DEPTO_STATE:  [MessageHandler(filters.TEXT & ~filters.COMMAND, get_depto)],
        },
        fallbacks=[CommandHandler("start", start)],
    )

    app.add_handler(conv)
    app.add_handler(CommandHandler("excel",   send_excel))
    app.add_handler(CommandHandler("resumen", resumen))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    logger.info("Bot COVISIAN iniciado.")
    app.run_polling()

if __name__ == "__main__":
    main()
